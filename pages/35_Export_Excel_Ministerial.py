"""
AGROVISION — Export Excel Multi-Sheet Ministerial
Ziua 35 | Autor: Prof. Asoc. Dr. Oliviu Mihnea Gamulescu

Scop:
    Generare fisier Excel cu mai multe sheet-uri pentru
    raportare APIA Gorj → APIA Central / Minister.

    Contine:
    - Sheet 1: Sumar judetean (KPI-uri PAC)
    - Sheet 2: Detalii per UAT
    - Sheet 3: Statistici per cultura
    - Sheet 4: Cereri neconforme
    - Sheet 5: Date complete (baza de date)

    Format profesional: culori, borduri, formatare numerica,
    antet APIA, gata de trimis la minister.
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

# ─── CONFIGURARE PAGINA ───────────────────────────────────────────────────────
st.set_page_config(
    page_title="Export Excel Ministerial | AGROVISION",
    page_icon="XLS",
    layout="wide"
)

# ─── DATE SIMULATE (aceleasi ca Ziua 34) ─────────────────────────────────────
UAT_GORJ = [
    {"uat": "Targu Jiu",     "lat": 45.042, "lon": 23.272},
    {"uat": "Rovinari",      "lat": 44.918, "lon": 23.165},
    {"uat": "Motru",         "lat": 44.807, "lon": 22.988},
    {"uat": "Bumbesti-Jiu",  "lat": 45.182, "lon": 23.391},
    {"uat": "Novaci",        "lat": 45.301, "lon": 23.673},
    {"uat": "Turceni",       "lat": 44.873, "lon": 23.401},
    {"uat": "Aninoasa",      "lat": 45.087, "lon": 23.522},
    {"uat": "Balesti",       "lat": 44.971, "lon": 23.489},
    {"uat": "Carbunesti",    "lat": 44.956, "lon": 23.527},
    {"uat": "Tismana",       "lat": 45.033, "lon": 22.992},
    {"uat": "Sacelele",      "lat": 45.118, "lon": 23.197},
    {"uat": "Vladuleni",     "lat": 44.842, "lon": 23.317},
]

CULTURI = ["grau", "porumb", "rapita", "floarea_soarelui",
           "lucerna", "pasune", "orz", "triticale"]

PLATII_BAZA = {
    "grau": 185, "porumb": 175, "rapita": 190,
    "floarea_soarelui": 172, "lucerna": 160,
    "pasune": 110, "orz": 178, "triticale": 165,
}


@st.cache_data
def genereaza_date(an: int) -> pd.DataFrame:
    rand = np.random.RandomState(an)
    rows = []
    fermier_id = 1
    for uat_info in UAT_GORJ:
        for _ in range(rand.randint(40, 120)):
            cultura   = rand.choice(CULTURI, p=[0.25,0.22,0.08,0.07,0.12,0.15,0.07,0.04])
            suprafata = round(rand.uniform(0.5, 25.0), 2)
            ndvi      = round(rand.uniform(0.2, 0.85), 3)
            conform   = rand.random() > 0.18
            risc      = rand.choice([1, 2, 3], p=[0.55, 0.30, 0.15])
            plata_baza  = round(suprafata * PLATII_BAZA[cultura], 2)
            plata_verde = round(suprafata * rand.uniform(40, 65), 2) if conform else 0
            plata_totala = round(plata_baza + plata_verde if conform else plata_baza * 0.3, 2)
            rows.append({
                "fermier_id": f"GJ{an}-{fermier_id:05d}",
                "uat": uat_info["uat"],
                "cultura": cultura,
                "suprafata": suprafata,
                "ndvi": ndvi,
                "conform": conform,
                "risc": risc,
                "plata_baza": plata_baza,
                "plata_verde": plata_verde,
                "plata_totala": plata_totala,
                "an": an,
            })
            fermier_id += 1
    return pd.DataFrame(rows)


# ─── STILURI EXCEL ────────────────────────────────────────────────────────────

ALBASTRU_APIA  = "FF0052A5"
VERDE_CONF     = "FF28A745"
ROSU_NECONF    = "FFDC3545"
GALBEN_RISC    = "FFFFC107"
ALB            = "FFFFFFFF"
GRI_DESCHIS    = "FFF8F9FA"
GRI_HEADER     = "FFE9ECEF"

def stil_header(ws, row: int, cols: int, text: str,
                culoare_bg: str = ALBASTRU_APIA,
                culoare_text: str = ALB,
                font_size: int = 12):
    """Scrie un rand de titlu/header pe toata latimea."""
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=cols
    )
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = Font(bold=True, size=font_size, color=culoare_text)
    cell.fill      = PatternFill("solid", fgColor=culoare_bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def stil_col_header(ws, row: int, headers: list,
                    culoare_bg: str = GRI_HEADER):
    """Scrie randul de antete coloane cu stil."""
    border_thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = Font(bold=True, size=10)
        cell.fill      = PatternFill("solid", fgColor=culoare_bg)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border_thin
    ws.row_dimensions[row].height = 30


def scrie_rand_date(ws, row: int, valori: list,
                    bg: str = ALB, bold: bool = False):
    """Scrie un rand de date cu borduri."""
    border_thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col, val in enumerate(valori, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill   = PatternFill("solid", fgColor=bg)
        cell.font   = Font(size=9, bold=bold)
        cell.border = border_thin
        cell.alignment = Alignment(vertical="center")


def ajusteaza_coloane(ws, latime_min: int = 10, latime_max: int = 35):
    """Ajusteaza latimea coloanelor dupa continut."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, latime_min), latime_max
        )


# ─── GENERARE EXCEL ───────────────────────────────────────────────────────────

def genereaza_excel_ministerial(df: pd.DataFrame, an: int,
                                 inspector: str) -> bytes:
    """
    Genereaza fisier Excel cu 5 sheet-uri pentru raportare ministerial APIA.
    """
    wb = Workbook()
    wb.remove(wb.active)   # sterge sheet-ul implicit

    # ── SHEET 1: SUMAR JUDETEAN ────────────────────────────────────────────────
    ws1 = wb.create_sheet("1. Sumar Judetean")

    total_fermieri  = len(df)
    sup_totala      = df["suprafata"].sum()
    plata_totala    = df["plata_totala"].sum()
    rata_conf       = df["conform"].mean() * 100
    nr_neconf       = (~df["conform"]).sum()
    plata_medie_ha  = plata_totala / sup_totala if sup_totala > 0 else 0

    stil_header(ws1, 1, 3,
                "AGENTIA DE PLATI SI INTERVENTIE PENTRU AGRICULTURA",
                font_size=13)
    stil_header(ws1, 2, 3,
                f"Centrul Judetean Gorj — Raport PAC Campania {an}",
                culoare_bg=GRI_HEADER, culoare_text="FF000000", font_size=11)
    stil_header(ws1, 3, 3,
                f"Intocmit: {date.today().strftime('%d.%m.%Y')} | Inspector: {inspector}",
                culoare_bg=ALB, culoare_text="FF666666", font_size=9)

    ws1.append([])
    stil_col_header(ws1, 5, ["Indicator", "Valoare", "UM"])

    indicatori = [
        ("Fermieri inregistrati",     total_fermieri,             "nr."),
        ("Suprafata totala declarata", round(sup_totala, 2),       "ha"),
        ("Plati totale PAC",          round(plata_totala, 2),     "EUR"),
        ("Plata medie per ha",        round(plata_medie_ha, 2),   "EUR/ha"),
        ("Cereri CONFORME",           total_fermieri - nr_neconf, "nr."),
        ("Cereri NECONFORME",         nr_neconf,                  "nr."),
        ("Rata conformitate",         round(rata_conf, 1),        "%"),
        ("UAT-uri acoperite",         df["uat"].nunique(),        "nr."),
        ("Culturi declarate",         df["cultura"].nunique(),    "nr."),
    ]

    for i, (ind, val, um) in enumerate(indicatori):
        bg = GRI_DESCHIS if i % 2 == 0 else ALB
        # Colorare speciala pentru rata conformitate
        if ind == "Cereri NECONFORME":
            bg = "FFFFD7D7"
        elif ind == "Rata conformitate":
            bg = "FFD4EDDA" if rata_conf >= 80 else "FFFFD7D7"
        scrie_rand_date(ws1, 6 + i, [ind, val, um], bg=bg)

    ajusteaza_coloane(ws1)

    # ── SHEET 2: DETALII UAT ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("2. Detalii per UAT")

    stil_header(ws2, 1, 7, f"Statistici per UAT — Campania PAC {an} | Gorj")

    df_uat = (
        df.groupby("uat")
        .agg(
            fermieri=("fermier_id", "count"),
            suprafata=("suprafata", "sum"),
            plata_baza=("plata_baza", "sum"),
            plata_verde=("plata_verde", "sum"),
            plata_totala=("plata_totala", "sum"),
            neconforme=("conform", lambda x: (~x).sum()),
            conform_pct=("conform", "mean"),
        )
        .reset_index()
        .sort_values("suprafata", ascending=False)
    )
    df_uat["conform_pct"] = (df_uat["conform_pct"] * 100).round(1)

    stil_col_header(ws2, 2, [
        "UAT", "Fermieri", "Suprafata (ha)",
        "Plata baza (EUR)", "Plata verde (EUR)",
        "Plata totala (EUR)", "Neconforme", "Conform (%)"
    ])

    for i, row in enumerate(df_uat.itertuples(index=False)):
        bg = GRI_DESCHIS if i % 2 == 0 else ALB
        # Rosu daca conformitate < 75%
        if row.conform_pct < 75:
            bg = "FFFFD7D7"
        scrie_rand_date(ws2, 3 + i, [
            row.uat,
            row.fermieri,
            round(row.suprafata, 1),
            round(row.plata_baza, 0),
            round(row.plata_verde, 0),
            round(row.plata_totala, 0),
            row.neconforme,
            f"{row.conform_pct:.1f}%"
        ], bg=bg)

    # Rand total
    scrie_rand_date(ws2, 3 + len(df_uat), [
        "TOTAL JUDET",
        df_uat["fermieri"].sum(),
        round(df_uat["suprafata"].sum(), 1),
        round(df_uat["plata_baza"].sum(), 0),
        round(df_uat["plata_verde"].sum(), 0),
        round(df_uat["plata_totala"].sum(), 0),
        df_uat["neconforme"].sum(),
        f"{df['conform'].mean()*100:.1f}%"
    ], bg="FFE8F4FD", bold=True)

    ajusteaza_coloane(ws2)

    # Grafic bare in ws2
    chart = BarChart()
    chart.type    = "col"
    chart.title   = f"Suprafata pe UAT — Campania {an}"
    chart.y_axis.title = "Suprafata (ha)"
    chart.x_axis.title = "UAT"
    chart.style   = 10
    chart.width   = 18
    chart.height  = 12

    data_ref = Reference(
        ws2,
        min_col=3, max_col=3,
        min_row=2, max_row=2 + len(df_uat)
    )
    cats_ref = Reference(ws2, min_col=1, min_row=3, max_row=2 + len(df_uat))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws2.add_chart(chart, f"A{5 + len(df_uat)}")

    # ── SHEET 3: STATISTICI CULTURI ───────────────────────────────────────────
    ws3 = wb.create_sheet("3. Statistici Culturi")

    stil_header(ws3, 1, 6, f"Statistici per Cultura — Campania PAC {an} | Gorj")

    df_cult = (
        df.groupby("cultura")
        .agg(
            fermieri=("fermier_id", "count"),
            suprafata=("suprafata", "sum"),
            ndvi_mediu=("ndvi", "mean"),
            plata_totala=("plata_totala", "sum"),
            plata_ha=("plata_totala", lambda x:
                      x.sum() / df.loc[x.index, "suprafata"].sum()),
        )
        .reset_index()
        .sort_values("suprafata", ascending=False)
    )

    stil_col_header(ws3, 2, [
        "Cultura", "Fermieri", "Suprafata (ha)",
        "NDVI mediu", "Plata totala (EUR)", "Plata/ha (EUR)"
    ])

    culori_culturi = {
        "grau":              "FFFFF3CD",
        "porumb":            "FFFFF9E6",
        "rapita":            "FFFFE8CC",
        "floarea_soarelui":  "FFFFF0CC",
        "lucerna":           "FFD4EDDA",
        "pasune":            "FFE8F5E9",
        "orz":               "FFFFF8E1",
        "triticale":         "FFFDE8E8",
    }

    for i, row in enumerate(df_cult.itertuples(index=False)):
        bg = culori_culturi.get(row.cultura, ALB)
        scrie_rand_date(ws3, 3 + i, [
            row.cultura,
            row.fermieri,
            round(row.suprafata, 1),
            round(row.ndvi_mediu, 3),
            round(row.plata_totala, 0),
            round(row.plata_ha, 2),
        ], bg=bg)

    # Rand total
    scrie_rand_date(ws3, 3 + len(df_cult), [
        "TOTAL",
        df_cult["fermieri"].sum(),
        round(df_cult["suprafata"].sum(), 1),
        round(df["ndvi"].mean(), 3),
        round(df_cult["plata_totala"].sum(), 0),
        round(df["plata_totala"].sum() / df["suprafata"].sum(), 2),
    ], bg="FFE8F4FD", bold=True)

    ajusteaza_coloane(ws3)

    # Grafic pie culturi
    pie = PieChart()
    pie.title  = f"Distributie suprafata pe culturi — {an}"
    pie.style  = 10
    pie.width  = 15
    pie.height = 12

    labels = Reference(ws3, min_col=1, min_row=3, max_row=2 + len(df_cult))
    data   = Reference(ws3, min_col=3, min_row=2, max_row=2 + len(df_cult))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws3.add_chart(pie, f"A{5 + len(df_cult)}")

    # ── SHEET 4: CERERI NECONFORME ────────────────────────────────────────────
    ws4 = wb.create_sheet("4. Cereri Neconforme")

    stil_header(ws4, 1, 7,
                f"Cereri NECONFORME — Campania {an} | Gorj",
                culoare_bg=ROSU_NECONF)

    df_nc = df[~df["conform"]].sort_values(
        ["risc", "suprafata"], ascending=[False, False]
    ).reset_index(drop=True)

    risc_label = {1: "SCAZUT", 2: "MEDIU", 3: "RIDICAT"}
    df_nc["risc_txt"] = df_nc["risc"].map(risc_label)

    stil_col_header(ws4, 2, [
        "ID Fermier", "UAT", "Cultura",
        "Suprafata (ha)", "NDVI", "Risc PAC", "Plata (EUR)"
    ])

    for i, row in enumerate(df_nc.itertuples(index=False)):
        if row.risc == 3:
            bg = "FFFFD7D7"   # rosu deschis — risc ridicat
        elif row.risc == 2:
            bg = "FFFFF3CD"   # galben — risc mediu
        else:
            bg = ALB
        scrie_rand_date(ws4, 3 + i, [
            row.fermier_id,
            row.uat,
            row.cultura,
            row.suprafata,
            row.ndvi,
            row.risc_txt,
            round(row.plata_totala, 2),
        ], bg=bg)

    # Sumar neconforme
    nr_row = 3 + len(df_nc) + 1
    stil_header(ws4, nr_row, 7,
                f"Total neconforme: {len(df_nc)} cereri | "
                f"Risc ridicat: {(df_nc['risc']==3).sum()} | "
                f"Plati afectate: {df_nc['plata_totala'].sum():,.0f} EUR",
                culoare_bg=GALBEN_RISC, culoare_text="FF000000", font_size=10)

    ajusteaza_coloane(ws4)

    # ── SHEET 5: DATE COMPLETE ────────────────────────────────────────────────
    ws5 = wb.create_sheet("5. Date Complete")

    stil_header(ws5, 1, len(df.columns),
                f"Date Complete Campania {an} — export baza de date",
                culoare_bg=GRI_HEADER, culoare_text="FF000000")

    stil_col_header(ws5, 2, list(df.columns))

    for i, row in enumerate(df.itertuples(index=False)):
        bg = GRI_DESCHIS if i % 2 == 0 else ALB
        scrie_rand_date(ws5, 3 + i, list(row), bg=bg)

    ajusteaza_coloane(ws5, latime_min=8, latime_max=25)

    # Salveaza in buffer
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# INTERFATA STREAMLIT
# ═══════════════════════════════════════════════════════════════════════════════

st.title("Ziua 35 - Export Excel Multi-Sheet Ministerial")
st.markdown(
    "**Fisier Excel profesional** cu 5 sheet-uri pentru raportare "
    "APIA Gorj → APIA Central / Minister. Culori, borduri, grafice, formatare numerica."
)

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("Configurare Export")

    an_ales = st.selectbox("Campania", [2024, 2023, 2022])
    inspector = st.text_input(
        "Inspector",
        value="Prof. Asoc. Dr. Oliviu Mihnea Gamulescu"
    )

    st.divider()
    st.markdown("**Sheet-uri generate:**")
    for s in ["1. Sumar Judetean", "2. Detalii per UAT",
              "3. Statistici Culturi", "4. Cereri Neconforme",
              "5. Date Complete"]:
        st.markdown(f"- {s}")

    st.divider()
    st.caption("openpyxl — stiluri native Excel")
    st.caption("Nu necesita Excel instalat")

# ─── DATE ─────────────────────────────────────────────────────────────────────
df = genereaza_date(an_ales)

# ─── TABS ─────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs([
    "Previzualizare date",
    "Structura Excel",
    "Generare & Download",
    "Teorie openpyxl"
])


# ── TAB 1: PREVIZUALIZARE ─────────────────────────────────────────────────────
with tab1:
    st.subheader(f"Date campanie {an_ales} — Gorj")

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("Fermieri", len(df))
    with k2:
        st.metric("Suprafata", f"{df['suprafata'].sum():,.0f} ha")
    with k3:
        st.metric("Plati totale", f"{df['plata_totala'].sum():,.0f} EUR")
    with k4:
        st.metric("Conformitate", f"{df['conform'].mean()*100:.1f}%")

    st.divider()

    col_v1, col_v2 = st.columns(2)

    with col_v1:
        df_cult_prev = (
            df.groupby("cultura")["suprafata"]
            .sum().reset_index()
            .sort_values("suprafata", ascending=False)
        )
        fig = px.bar(df_cult_prev, x="cultura", y="suprafata",
                     title="Suprafata per cultura (ha)",
                     color="cultura",
                     color_discrete_sequence=px.colors.qualitative.Set2)
        fig.update_layout(height=320, showlegend=False)
        st.plotly_chart(fig, use_container_width=True)

    with col_v2:
        df_uat_prev = (
            df.groupby("uat")["plata_totala"]
            .sum().reset_index()
            .sort_values("plata_totala", ascending=False)
        )
        fig2 = px.bar(df_uat_prev, x="uat", y="plata_totala",
                      title="Plati totale per UAT (EUR)",
                      color_discrete_sequence=["#0052A5"])
        fig2.update_layout(height=320, xaxis_tickangle=-30)
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown(f"**Primele 10 randuri** din datele complete:")
    st.dataframe(df.head(10), use_container_width=True, hide_index=True)


# ── TAB 2: STRUCTURA EXCEL ────────────────────────────────────────────────────
with tab2:
    st.subheader("Structura fisierului Excel generat")

    for sheet, desc, culoare in [
        ("1. Sumar Judetean",
         "KPI-uri cheie: nr. fermieri, ha, EUR, conformitate. "
         "Antet APIA albastru, culori semaforizare conformitate.",
         "#e8f4fd"),
        ("2. Detalii per UAT",
         "Rand per UAT cu suprafata, plata baza/verde/totala, neconforme. "
         "Randuri rosii pentru UAT cu conformitate < 75%. Grafic bare inclus.",
         "#d4edda"),
        ("3. Statistici Culturi",
         "Rand per cultura cu ha, NDVI mediu, plata totala si plata/ha. "
         "Culori distincte per cultura. Grafic pie inclus.",
         "#fff3cd"),
        ("4. Cereri Neconforme",
         "Lista completa cereri NECONFORME, sortata dupa risc si suprafata. "
         "Rosu = risc ridicat, galben = risc mediu. Sumar total jos.",
         "#f8d7da"),
        ("5. Date Complete",
         "Export complet al bazei de date — toate campurile, "
         "toate randurile. Fundal alternant pentru lizibilitate.",
         "#f8f9fa"),
    ]:
        st.markdown(
            f'<div style="background:{culoare};border-radius:8px;'
            f'padding:12px 16px;margin:6px 0;border-left:4px solid #0052A5;">'
            f'<b>{sheet}</b><br><span style="font-size:13px;">{desc}</span>'
            f'</div>',
            unsafe_allow_html=True
        )


# ── TAB 3: GENERARE & DOWNLOAD ────────────────────────────────────────────────
with tab3:
    st.subheader("Generare Fisier Excel")

    st.markdown(f"""
    **Parametri export:**
    - Campania: **{an_ales}**
    - Inspector: **{inspector}**
    - Fermieri: **{len(df):,}**
    - Suprafata totala: **{df['suprafata'].sum():,.1f} ha**
    - Plati totale: **{df['plata_totala'].sum():,.0f} EUR**
    - Neconforme: **{(~df['conform']).sum()}** cereri
    """)

    st.divider()

    excel_bytes = genereaza_excel_ministerial(df, an_ales, inspector)

    st.download_button(
        label=f"Descarca Excel Ministerial — Campania {an_ales}",
        data=excel_bytes,
        file_name=f"Raport_PAC_Gorj_{an_ales}_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary"
    )

    st.success(
        "Fisierul Excel contine 5 sheet-uri cu formatare profesionala, "
        "culori semaforizare si grafice incluse. "
        "Deschide in Excel sau Google Sheets."
    )

    st.divider()

    # Comparatie 3 ani
    st.subheader("Export comparativ 3 ani (2022-2024)")
    if st.button("Genereaza Excel comparativ 2022-2024"):
        buf_comp = io.BytesIO()
        with pd.ExcelWriter(buf_comp, engine="openpyxl") as writer:
            for an in [2022, 2023, 2024]:
                df_an = genereaza_date(an)
                df_sumar = df_an.groupby("uat").agg(
                    fermieri=("fermier_id", "count"),
                    suprafata=("suprafata", "sum"),
                    plata_totala=("plata_totala", "sum"),
                    conform_pct=("conform", "mean"),
                ).reset_index()
                df_sumar["conform_pct"] = (df_sumar["conform_pct"] * 100).round(1)
                df_sumar.to_excel(writer, sheet_name=f"Campania {an}", index=False)

        st.download_button(
            label="Descarca Excel comparativ 3 ani",
            data=buf_comp.getvalue(),
            file_name=f"Comparativ_PAC_Gorj_2022_2024.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary"
        )


# ── TAB 4: TEORIE OPENPYXL ────────────────────────────────────────────────────
with tab4:
    st.subheader("openpyxl — Formatare Excel Profesionala")

    st.markdown("""
### Concepte cheie openpyxl

```python
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active

# Scriere celula
ws["A1"] = "Text"
ws.cell(row=1, column=1, value="Text")

# Font
cell.font = Font(bold=True, size=12, color="FF0052A5")
# color = ARGB hex (FF = opac, 0052A5 = albastru APIA)

# Fundal
cell.fill = PatternFill("solid", fgColor="FFD4EDDA")

# Aliniere
cell.alignment = Alignment(horizontal="center", vertical="center")

# Borduri
thin = Side(style="thin")
cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Merge celule
ws.merge_cells("A1:E1")

# Latime coloana
ws.column_dimensions["A"].width = 20

# Inaltime rand
ws.row_dimensions[1].height = 25
```

### Grafice in Excel cu openpyxl

```python
from openpyxl.chart import BarChart, PieChart, Reference

chart = BarChart()
chart.title = "Titlu grafic"
chart.style = 10        # stil vizual 1-48

# Datele pentru grafic
data = Reference(ws, min_col=2, min_row=1, max_row=10)
categories = Reference(ws, min_col=1, min_row=2, max_row=10)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Adauga graficul in sheet
ws.add_chart(chart, "D5")   # pozitie in sheet
```

### De ce openpyxl vs pandas to_excel?

| Functionalitate | pandas | openpyxl |
|----------------|--------|---------|
| Scriere simpla | ✅ rapid | ✅ rapid |
| Culori celule | ❌ | ✅ |
| Borduri | ❌ | ✅ |
| Merge celule | ❌ | ✅ |
| Grafice | ❌ | ✅ |
| Formule Excel | ❌ | ✅ |
| Control total format | ❌ | ✅ |

**Recomandare:** pandas pentru export rapid, openpyxl pentru documente profesionale.
    """)


# ─── FOOTER ───────────────────────────────────────────────────────────────────
st.divider()
st.caption(
    "Ziua 35 - AGROVISION | Export Excel Multi-Sheet Ministerial | "
    "openpyxl + stiluri profesionale APIA | "
    "Prof. Asoc. Dr. Oliviu Mihnea Gamulescu, UCB Targu Jiu 2026"
)
